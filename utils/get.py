import openpyxl
from utils.common import printLine

emails_index = 1
channels_index = 2


def channelNameColumnIndex(sheet):
    print("Searching for the 'Channel Name' column index")

    for row_index in range(1, 2):
        print(sheet.max_column)
        for column_index in range(1, sheet.max_column+1):
            cell = sheet.cell(row=row_index, column=column_index)
            value = cell.value
            print(value)

            if "Channel Name" == value:
                print("'Channel Name' found at column {}".format(column_index))
                printLine()
                channels_index = column_index
                return column_index
    
    print("COLUMN NOT FOUND!")
    printLine()
    return None
    

def emailsColumnIndex(sheet):
    print("Searching for the 'User E-Mail' column index")

    for row_index in range(1, 2):
        print(sheet.max_column)
        for column_index in range(1, sheet.max_column+1):
            cell = sheet.cell(row=row_index, column=column_index)
            value = cell.value
            print(value)

            if "User E-Mail" == value:
                print("'User E-Mail' found at column {}".format(column_index))
                printLine()
                emails_index = column_index
                return column_index
    
    print("COLUMN NOT FOUND!")
    printLine()
    return None


def emailsListFromCategory(sheet, channel, channel_row_index):
    emails_list = []

    for row_index in range(channel_row_index, sheet.max_row):
        #it jumps the first line because it's where the header is located
        channel_name = sheet.cell(row=row_index, column=channels_index).value
        email = sheet.cell(row=row_index, column=emails_index).value
        
        if channel == channel_name:
            emails_list.append(email)

    print("Emails list for {} filled successfully with ".format(channel)+
          "{} items".format(len(emails_list)))
    printLine()

    return emails_list


def channelsDict(sheet):
    # channels_dict = 
    # {
    #   "channel name": [ "email 1 ", "email 2", ... ] 
    # }

    channels_dict = {}
    past_email = ""

    print("Starting filling the channels dictionary")
    printLine()

    for row_index in range(2, sheet.max_row):
        # it jumps the first line because it's where the header is located

        channel_name = sheet.cell(row=row_index, column=channels_index).value
        email = sheet.cell(row=row_index, column=emails_index).value
        
        if channel_name != None and email != None:
            if email != past_email:
                print("Getting Channels list from {}".format(email) + "category")
                past_email = email
                channels_dict[channel_name] = emailsListFromCategory(sheet, channel_name, row_index)
    
    print("Finished filling the dictionary")
    printLine()
    return channels_dict